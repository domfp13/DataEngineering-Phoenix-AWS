# Adding other libs
import sys
import string
from openpyxl import load_workbook #xlsx
from xlrd import open_workbook #xls
import xlrd
from io import StringIO
import datetime

# Adding local classes
from etl.transforms import Ingram, Synnex, TechData, IngramCAD, CADTechDataOG, CADTechData, IngramMicroBMO

# Adding local functions
from etl.functions import generateExcelColumns, sortingList, customerNameLinearSearch
from etl.output import to_csv

# Getting data from DW
# sorted_by_value = sortingList(getDataFromTLKPOrderClientData({'name':'T_LKP_ORDER_CLIENT_DATA.csv', 'bucket':'dw_file_export'}))

def xlrdFuntionTime(excelCell, wb):
    # Check for if the value is empy
    poEta = ''

    if isinstance(excelCell, str):
        poEta = excelCell
    else:
        poEta = datetime.datetime(*xlrd.xldate_as_tuple(excelCell, wb.datemode)).strftime('%m/%d/%Y') 

    return poEta

def parse_CADTechData(file):
    try:
        # This gets the workbook and the tap
        wb = open_workbook(filename=file)
        ws = wb.sheet_by_index(0)

        columndic = generateExcelColumns(counterbegin=0, sizeOfColumns=1)
        data = []

        for row in range(4, ws.nrows):

            CADTechData_row = CADTechData(
               customerName =  ws.cell(row, columndic['E']).value,
               forecastName = ws.cell(row, columndic['Z']).value,
               mfgPartno = ws.cell(row, columndic['J']).value,
               reservedQuantity = ws.cell(row, columndic['O']).value,
               availQuantity = ws.cell(row, columndic['Q']).value,
               boQty = ws.cell(row, columndic['R']).value,
               age = ws.cell(row, columndic['N']).value,
               poEta = xlrdFuntionTime(excelCell = ws.cell(row, columndic['S']).value, wb = wb),
               oemPo = ws.cell(row, columndic['T']).value
            )
            
            if CADTechData_row.nonempty():
                # CADTechData_row.customerId = customerNameLinearSearch(CADTechData_row.customerName.upper(), sorted_by_value)
                data.append(CADTechData_row.__dict__)

            del CADTechData_row        
    except Exception as e:
        print(e)
    else:
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0001_TECHDATA_CAD_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, CADTechDataOG.keys())

def parse_CADTechDataOG(file):
    try:
        # This gets the workbook and the tap
        wb = open_workbook(filename=file)
        ws = wb.sheet_by_index(0)

        columndic = generateExcelColumns(counterbegin=0, sizeOfColumns=1)
        data = []

        for row in range(4, ws.nrows):

            CADTechDataOG_row = CADTechDataOG(
               customerName =  ws.cell(row, columndic['E']).value,
               forecastName = ws.cell(row, columndic['X']).value, #.encode('utf-8')
               mfgPartno = ws.cell(row, columndic['I']).value,
               reservedQuantity = ws.cell(row, columndic['N']).value,
               availQuantity = ws.cell(row, columndic['P']).value,
               boQty = ws.cell(row, columndic['Q']).value,
               age = ws.cell(row, columndic['M']).value,
               poEta = xlrdFuntionTime(excelCell = ws.cell(row, columndic['R']).value, wb = wb),
               oemPo = ws.cell(row, columndic['S']).value
            )
            
            if CADTechDataOG_row.nonempty():
                # CADTechDataOG_row.customerId = customerNameLinearSearch(CADTechDataOG_row.customerName.upper(), sorted_by_value)
                data.append(CADTechDataOG_row.__dict__)

            del CADTechDataOG_row

    except Exception as e:
        print(e)
    else:
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0002_TECHDATAOG_CAD_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, CADTechDataOG.keys())

def parse_Ingram_Micro_BMO(file):

    try:
        # This gets the workbook and the tap
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        columndic = generateExcelColumns(counterbegin=1, sizeOfColumns=1)
        data = []

        for row in range(2, ws.max_row+1):
            ingramMicroBMO_row = IngramMicroBMO(
                mfgPartno = ws.cell(row = row, column = columndic['B']).value, #C
                availQuantity = ws.cell(row = row, column = columndic['G']).value, #J
                boQty = ws.cell(row = row, column = columndic['D']).value #F
            )
            
            if ingramMicroBMO_row.nonempty():
                # ingramMicroBMO_row.customerId = customerNameLinearSearch(ingramMicroBMO_row.customerName.upper(), sorted_by_value)
                data.append(ingramMicroBMO_row.__dict__)
            
            del ingramMicroBMO_row

    except Exception as e:
        print(e)
    else:
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0003_INGRAM_MICRO_BMO_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, IngramMicroBMO.keys())

def parse_synnex(file):
    try:
        wb = open_workbook(filename=file)
        ws = wb.sheet_by_index(0)

        columndic = generateExcelColumns(counterbegin=0, sizeOfColumns=1)
        data = []

        for row in range(1, ws.nrows):
            synnex_row = Synnex(
               customerName =  ws.cell(row, columndic['I']).value,
               forecastName = ws.cell(row, columndic['H']).value,
               mfgPartno = ws.cell(row, columndic['F']).value,
               reservedQuantity = ws.cell(row, columndic['J']).value,
               availQuantity = ws.cell(row, columndic['O']).value,
               boQty = ws.cell(row, columndic['P']).value,
               age = ws.cell(row, columndic['S']).value,
               poEta = xlrdFuntionTime(excelCell = ws.cell(row, columndic['T']).value, wb = wb),
               oemPo = ws.cell(row, columndic['U']).value
            )

            if synnex_row.nonempty():
                # synnex_row.customerId = customerNameLinearSearch(synnex_row.customerName.upper(), sorted_by_value)
                data.append(synnex_row.__dict__)
            
            del synnex_row

    except Exception as e:
        print(e)
    else:
        # After reading the wb it manages the memory
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0004_SYNNEX_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, Synnex.keys())

def parse_ingram(file):

    try:
        # This gets the workbook and the tap
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        columndic = generateExcelColumns(counterbegin=1, sizeOfColumns=1)
        data = []

        for row in range(2, ws.max_row+1):
            ingram_row = Ingram(
                customerName = '', #ws.cell(row = row, column = columndic['K']).value,
                mfgPartno = ws.cell(row = row, column = columndic['C']).value,
                reservedQuantity = ws.cell(row = row, column = columndic['H']).value,
                availQuantity = ws.cell(row = row, column = columndic['J']).value
            )
            
            if ingram_row.nonempty():
                # ingram_row.customerId = customerNameLinearSearch(ingram_row.customerName.upper(), sorted_by_value)
                data.append(ingram_row.__dict__)
            
            del ingram_row

    except Exception as e:
        print(e)
    else:
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0005_INGRAM_MICRO_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, Ingram.keys())

def parse_techdata(file):
    try:
        
        # This gets the workbook and the tap
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        columndic = generateExcelColumns(counterbegin=1, sizeOfColumns=2)
        data = []

        for row in range(4, ws.max_row+1):
            
            techdata_row = TechData(
                customerName = ws.cell(row = row, column = columndic['I']).value,
                forecastName = ws.cell(row = row, column = columndic['I']).value, #This has to be parse out
                mfgPartno = ws.cell(row = row, column = columndic['N']).value,
                availQuantity = ws.cell(row = row, column = columndic['Z']).value,
                boQty = ws.cell(row = row, column = columndic['AA']).value
            )

            if techdata_row.nonempty():
                # techdata_row.customerId = customerNameLinearSearch(techdata_row.customerName.upper(), sorted_by_value)
                data.append(techdata_row.__dict__)

            del techdata_row

    except Exception as e:
        print(e)
    else:
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0007_TECHDATA_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, TechData.keys())

def parse_ingramCAD(file):
    try:
        # This gets the workbook and the tap
        wb = open_workbook(filename=file)
        ws = wb.sheet_by_index(0)

        columndic = generateExcelColumns(counterbegin=0, sizeOfColumns=1)
        data = []

        for row in range(2, ws.nrows):

            ingramCAD_row = IngramCAD(
               customerName =  ws.cell(row, columndic['B']).value,
               mfgPartno = ws.cell(row, columndic['D']).value,
               availQuantity = ws.cell(row, columndic['H']).value      
            )

            if ingramCAD_row.nonempty():
                # ingramCAD_row.customerId = customerNameLinearSearch(ingramCAD_row.customerName.upper(), sorted_by_value)
                data.append(ingramCAD_row.__dict__)

            del ingramCAD_row

    except Exception as e:
        print(e)
    else:
        # After reading the wb it manages the memory
        ws = None
        wb = None
        fileNameOutput = 'DBI_LOAD_0008_CAD_INGRAM_MICRO_'+(str(datetime.date.today()))+'.csv'
        to_csv(fileNameOutput, data, IngramCAD.keys())